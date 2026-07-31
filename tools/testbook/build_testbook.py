#!/usr/bin/env python3
"""
Build the SIMF Production-Readiness Test Book (Excel) from the live catalogue.

Source of truth stays in the Markdown docs; this script only PROJECTS them into
a human-runnable .xlsx so the workbook never drifts by hand-copy. Re-run after
the catalogue or the business-flows doc changes.

Inputs
  docs/tests/e2e/*.md              per-page Gherkin catalogue (2,100+ scenarios)
  docs/tests/e2e/README.md         used to assign a Module/Area to each file
  docs/tests/SIMF-Business-Flows.md  the 15 cross-page journeys (optional until authored)

Output
  docs/tests/SIMF-Production-Readiness-TestBook.xlsx

It extracts ONLY coverage-matrix rows (id / scenario / type / priority) — never a
Gherkin body from a per-page file, and never a front-matter line — so no literal
secret from an Auth-setup line can leak into the workbook.

Catalogue size
  Every run re-derives the catalogue size (page count + Coverage-matrix scenario
  count) and EMITS it in three places, so no document has to hand-copy a number
  that goes stale:
    1. stdout — the "CATALOGUE SIZE" block at the end of the run;
    2. the workbook, sheet "00 · How to use" — a dated provenance line;
    3. the workbook, sheet "01 · Summary" — a "Catalogue size" block under the
       roll-up totals.
  It also compares the derived figures against the roll-up in
  docs/tests/e2e/README.md and prints a MISMATCH warning if they disagree. The
  warning is not fatal — the workbook must still regenerate so the number CAN be
  fixed. The build gate is
  tests/SIMF.ControlPanel.Tests/E2eCatalogueIntegrityTests
  .The_index_roll_up_matches_the_catalogue_it_describes, which fails the build on
  the same mismatch. Prose that needs a count should cite the README roll-up (or
  re-run this script), never a literal copied into a paragraph.
"""

from __future__ import annotations

import datetime
import os
import re
import sys
import glob

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
E2E_DIR = os.path.join(REPO, "docs", "tests", "e2e")
README = os.path.join(E2E_DIR, "README.md")
FLOWS_DOC = os.path.join(REPO, "docs", "tests", "SIMF-Business-Flows.md")
OUT = os.path.join(REPO, "docs", "tests", "SIMF-Production-Readiness-TestBook.xlsx")

SKIP_FILES = {"_TEMPLATE.md", "README.md", "E2E-TEST-PLAN.md"}
# Namespace may itself contain a hyphen (per-page ids like E2E-HAL-001 AND
# two-segment business-flow ids like E2E-BF-01-001), so allow hyphens in the
# namespace portion before the final zero-padded sequence number.
ID_RE = re.compile(r"^E2E-[A-Z0-9][A-Z0-9-]*-\d{3,4}$")
SCEN_ID_RE = re.compile(r"E2E-[A-Z0-9][A-Z0-9-]*-\d{3,4}")
# A Coverage-matrix ROW — the id is the first cell. Deliberately identical to
# E2eCatalogueIntegrityTests.MatrixRow so the size this script emits is the same
# size the build gate enforces. Narrower than extract_scenarios() below, which is
# tolerant on purpose (it also harvests ids from `### E2E-… — title` headers and
# from inverted tables) because a workbook row is worth having even when a page
# authored its matrix unusually. The two numbers answer different questions:
# "how big is the catalogue" vs "how many rows did the projector build".
MATRIX_ROW_RE = re.compile(r"^\|\s*(E2E-[A-Z0-9][A-Z0-9-]*-\d{3,4})\s*\|")

# ----- column layout (shared by every test sheet) --------------------------
COLS = [
    ("ID", 16),
    ("Surface", 12),
    ("Area / Flow", 26),
    ("Page / Route", 26),
    ("Test (scenario)", 60),
    ("Steps / Gherkin", 70),
    ("Test data", 26),
    ("Expected result", 34),
    ("Priority", 9),
    ("Type", 12),
    ("Catalogue ref", 26),
    ("Comment", 24),
    ("Time (min)", 10),
    ("Status", 12),
    ("Developer comment", 28),
    ("Evidence", 22),
]
STATUS_COL = 14  # 'Status' (1-based)
STATUS_LETTER = get_column_letter(STATUS_COL)
STATUS_CHOICES = ["Not Run", "Pass", "Fail", "Blocked", "N/A"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SUB_FONT = Font(italic=True, size=10, color="404040")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
BLOCK_FILL = PatternFill("solid", fgColor="FFEB9C")
NA_FILL = PatternFill("solid", fgColor="E7E6E6")


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
def read(path: str) -> str:
    with open(path, "r", encoding="utf-8") as fh:
        return fh.read()


def split_row(line: str):
    """Split a markdown table row into trimmed cells."""
    cells = line.strip().strip("|").split("|")
    return [c.strip() for c in cells]


def parse_readme_modules() -> dict:
    """Map each catalogue file name -> the README section heading it sits under."""
    modules = {}
    if not os.path.exists(README):
        return modules
    current = ""
    for line in read(README).splitlines():
        h = re.match(r"^#{2,4}\s+(.*)", line)
        if h:
            current = h.group(1).strip()
            continue
        for m in re.finditer(r"\[`([a-z0-9\-]+\.md)`\]", line):
            modules.setdefault(m.group(1), current)
    return modules


def parse_frontmatter(text: str) -> dict:
    """Pull Route / Surface / Pages covered from the leading `| | |` table."""
    info = {"route": "", "surface": ""}
    for line in text.splitlines()[:40]:
        cells = split_row(line) if line.strip().startswith("|") else []
        if len(cells) >= 2:
            key = cells[0].strip("* ").lower()
            val = cells[1]
            if key == "route" and not info["route"]:
                info["route"] = val.strip("` ")
            elif key in ("surface",) and not info["surface"]:
                info["surface"] = val
            elif key in ("pages covered", "page(s)") and not info["route"]:
                info["route"] = re.sub(r"[`\[\]]", "", val)[:60]
    return info


def find_matrix(text: str):
    """Return (header_cells, [data_rows]) for the first standard coverage matrix
    table (ID in the first column). Used by the business-flows parser, which
    authors a standard `| ID | Scenario | Type | Priority |` matrix."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if line.strip().lower().startswith("| id ") or re.match(
            r"^\|\s*id\s*\|", line.strip(), re.I
        ):
            header = [c.lower() for c in split_row(line)]
            rows = []
            j = i + 1
            if j < len(lines) and set(lines[j].replace("|", "").strip()) <= set("-: "):
                j += 1  # skip the |---|---| separator
            while j < len(lines) and lines[j].strip().startswith("|"):
                cells = split_row(lines[j])
                if cells and ID_RE.match(cells[0]):
                    rows.append(cells)
                j += 1
            if rows:
                return header, rows
    return None, []


NON_SCEN_HEADERS = {"type", "priority", "status", "page(s)", "phase", "id",
                    "scenario id", "test id"}


def extract_scenarios(text: str) -> dict:
    """Tolerant harvest of every scenario in a catalogue file, keyed by id.

    Handles all three real shapes: (a) a matrix with ID in column 0, (b) an
    inverted matrix like `| Function | Scenario id |` (ID in another column),
    and (c) files with NO matrix where scenarios are only `### E2E-… — title`
    headers. Merges: table data wins for type/priority; headers fill the title.
    """
    results: dict = {}

    # (c) scenario headers "## / ### E2E-XXX-NNN — title"
    for m in re.finditer(
        r"(?m)^#{2,4}\s+(E2E-[A-Z0-9]+-\d{3,4})\s*[—\-–]\s*(.+?)\s*$", text
    ):
        results.setdefault(m.group(1), {})["scenario"] = m.group(2).strip()

    # (a)/(b) any markdown table that carries an id cell
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        is_sep = (
            i + 1 < len(lines)
            and lines[i + 1].strip().startswith("|")
            and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
        )
        if line.strip().startswith("|") and is_sep:
            header = [c.lower() for c in split_row(line)]
            j = i + 2
            data = []
            while j < len(lines) and lines[j].strip().startswith("|"):
                data.append(split_row(lines[j]))
                j += 1
            # which column holds ids?
            idcol = -1
            for ci in range(len(header)):
                if any(ci < len(r) and ID_RE.match(r[ci]) for r in data):
                    idcol = ci
                    break
            if idcol >= 0:
                scencol = -1
                for want in ("scenario", "function", "description", "name"):
                    if want in header and header.index(want) != idcol:
                        scencol = header.index(want)
                        break
                if scencol < 0:
                    cand = [ci for ci in range(len(header))
                            if ci != idcol and header[ci] not in NON_SCEN_HEADERS]
                    scencol = cand[0] if cand else -1
                tcol = header.index("type") if "type" in header else -1
                pcol = header.index("priority") if "priority" in header else -1
                for r in data:
                    if idcol < len(r) and ID_RE.match(r[idcol]):
                        d = results.setdefault(r[idcol], {})
                        if scencol >= 0 and scencol < len(r) and r[scencol] and not d.get("scenario"):
                            d["scenario"] = r[scencol]
                        if tcol >= 0 and tcol < len(r) and r[tcol]:
                            d.setdefault("type", r[tcol])
                        if pcol >= 0 and pcol < len(r) and r[pcol]:
                            d.setdefault("priority", r[pcol])
            i = j
        else:
            i += 1
    return results


def id_sort_key(scenario_id: str):
    m = re.match(r"E2E-([A-Z0-9]+)-(\d+)", scenario_id)
    return (m.group(1), int(m.group(2))) if m else (scenario_id, 0)


def col_index(header, *names):
    for n in names:
        if n in header:
            return header.index(n)
    return -1


def surface_of(fname: str) -> str:
    if fname.startswith("cp-"):
        return "Control Panel"
    if fname.startswith("web-"):
        return "Website"
    if fname.startswith("mobile-"):
        return "Mobile App"
    return ""


def parse_catalogue_file(path: str, modules: dict):
    fname = os.path.basename(path)
    text = read(path)
    fm = parse_frontmatter(text)
    found = extract_scenarios(text)
    out = []
    for sid in sorted(found, key=id_sort_key):
        d = found[sid]
        out.append(
            {
                "id": sid,
                "surface": surface_of(fname) or fm["surface"],
                "area": modules.get(fname, ""),
                "route": fm["route"],
                "scenario": d.get("scenario", ""),
                "type": d.get("type", ""),
                "priority": (d.get("priority") or "P1"),
                "ref": f"e2e/{fname}",
                "steps": "",
            }
        )
    return out


def catalogue_files() -> list:
    """Every authored per-page catalogue file — index/template/playbook excluded."""
    return [
        path
        for path in sorted(glob.glob(os.path.join(E2E_DIR, "*.md")))
        if os.path.basename(path) not in SKIP_FILES
    ]


# Reuses surface_of() rather than re-testing the cp-/web-/mobile- prefixes, so
# the naming rule lives in exactly one place.
SURFACE_BUCKET = {"Control Panel": "cp", "Website": "web", "Mobile App": "mobile"}


def catalogue_size() -> dict:
    """Re-derive the catalogue size from the files themselves.

    Returns page/scenario totals plus the per-surface page split, so a document
    that needs a number can quote a generated figure instead of a hand-copied
    one. Counts Coverage-matrix rows only (see MATRIX_ROW_RE).
    """
    size = {"pages": 0, "scenarios": 0, "cp": 0, "web": 0, "mobile": 0, "other": 0}
    for path in catalogue_files():
        name = os.path.basename(path)
        size["pages"] += 1
        size[SURFACE_BUCKET.get(surface_of(name), "other")] += 1
        for line in read(path).splitlines():
            if MATRIX_ROW_RE.match(line):
                size["scenarios"] += 1
    return size


def readme_rollup() -> tuple:
    """The (pages, scenarios) the index claims, or (None, None) if unparseable."""
    if not os.path.exists(README):
        return (None, None)
    m = re.search(
        r"Pages catalogued:\*\*\s*(\d+)\b.*?Total scenarios:\*\*\s*(\d+)\b",
        read(README),
        re.S,
    )
    if not m:
        return (None, None)
    return (int(m.group(1)), int(m.group(2)))


def report_catalogue_size(size: dict) -> None:
    """Print the derived size and flag a stale index roll-up."""
    print("")
    print("CATALOGUE SIZE (re-derived this run — quote this, never a copied literal)")
    print(f"  pages catalogued:  {size['pages']}"
          f"  ({size['cp']} Control Panel + {size['mobile']} mobile"
          f" + {size['web']} Website + {size['other']} other)")
    print(f"  scenarios:         {size['scenarios']} Coverage-matrix rows")

    claimed_pages, claimed_scenarios = readme_rollup()
    if claimed_pages is None:
        print("  ! docs/tests/e2e/README.md has no parseable roll-up "
              "('**Pages catalogued:** N' / '**Total scenarios:** N') — cannot "
              "cross-check.", file=sys.stderr)
        return
    if (claimed_pages, claimed_scenarios) != (size["pages"], size["scenarios"]):
        print(f"  ! MISMATCH: docs/tests/e2e/README.md claims "
              f"{claimed_pages} pages / {claimed_scenarios} scenarios. Update the "
              f"roll-up — E2eCatalogueIntegrityTests"
              f".The_index_roll_up_matches_the_catalogue_it_describes fails the "
              f"build on this.", file=sys.stderr)
    else:
        print("  index roll-up agrees.")


def anchor_of(heading: str) -> str:
    a = heading.strip().lower()
    a = re.sub(r"[^\w\s-]", "", a)
    a = re.sub(r"\s+", "-", a)
    return a


def parse_business_flows(path: str):
    """Parse the assembled business-flows doc into per-scenario rows with Gherkin."""
    if not os.path.exists(path):
        return []
    text = read(path)
    lines = text.splitlines()
    out = []
    # locate each "## BF-xx — title" section
    sec_idx = [
        (i, re.match(r"^##\s+(BF-\d+)\s+[—-]\s+(.*)", ln))
        for i, ln in enumerate(lines)
    ]
    sections = [(i, m) for i, m in sec_idx if m]
    for k, (start, m) in enumerate(sections):
        flow_id, flow_title = m.group(1), m.group(2).strip()
        end = sections[k + 1][0] if k + 1 < len(sections) else len(lines)
        block = "\n".join(lines[start:end])
        # surface(s): scan intro for the words
        surfaces = []
        intro = block.lower()
        for key, name in (
            ("control panel", "CP"),
            ("website", "Web"),
            ("mobile", "App"),
            ("app ", "App"),
        ):
            if key in intro and name not in surfaces:
                surfaces.append(name)
        surface = "/".join(surfaces) if surfaces else "Cross"
        header, rows = find_matrix(block)
        i_scen = col_index(header or [], "scenario")
        i_type = col_index(header or [], "type")
        i_prio = col_index(header or [], "priority")
        # gherkin blocks in order
        gherkins = re.findall(r"```gherkin\s*\n(.*?)```", block, re.S)
        matrix_ids = [c[0] for c in rows]
        zipped = len(gherkins) == len(rows) and len(rows) > 0
        for idx, cells in enumerate(rows):
            steps = gherkins[idx].strip() if zipped else ""
            out.append(
                {
                    "id": cells[0],
                    "surface": surface,
                    "area": f"{flow_id} {flow_title}",
                    "route": "",
                    "scenario": cells[i_scen] if 0 <= i_scen < len(cells) else "",
                    "type": cells[i_type] if 0 <= i_type < len(cells) else "",
                    "priority": (cells[i_prio] if 0 <= i_prio < len(cells) else "P1") or "P1",
                    "ref": f"SIMF-Business-Flows.md#{anchor_of(flow_id + '--' + flow_title)}",
                    "steps": steps,
                    "flow_id": flow_id,
                }
            )
        if rows and not zipped:
            print(f"  ! {flow_id}: matrix rows ({len(rows)}) != gherkin blocks "
                  f"({len(gherkins)}) — steps left blank, see the doc", file=sys.stderr)
    return out


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------
def style_header(ws):
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c, value=COLS[c - 1][0])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = COLS[c - 1][1]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26


def write_rows(ws, rows):
    style_header(ws)
    r = 2
    for row in rows:
        vals = [
            row["id"], row["surface"], row["area"], row["route"], row["scenario"],
            row.get("steps", ""), row.get("data", ""), row.get("expected", ""),
            row["priority"], row["type"], "", "", "", "Not Run", "", "",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP if c in (3, 4, 5, 6, 7, 8, 12, 15) else CENTER if c in (2, 9, 10, 13, 14) else Alignment(vertical="top")
            cell.border = BORDER
        # catalogue ref hyperlink (col 11)
        ref = row.get("ref", "")
        if ref:
            cref = ws.cell(row=r, column=11, value=ref.split("#")[0].split("/")[-1])
            cref.hyperlink = ref
            cref.font = Font(color="0563C1", underline="single", size=10)
            cref.alignment = WRAP
        r += 1
    last = r - 1
    if last >= 2:
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS_CHOICES), allow_blank=True)
        dv.add(f"{STATUS_LETTER}2:{STATUS_LETTER}{last}")
        ws.add_data_validation(dv)
        rng = f"{STATUS_LETTER}2:{STATUS_LETTER}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N/A"'], fill=NA_FILL))
    return last


def build():
    modules = parse_readme_modules()

    size = catalogue_size()

    per_page = []
    for path in catalogue_files():
        per_page.extend(parse_catalogue_file(path, modules))

    cp = [r for r in per_page if r["surface"] == "Control Panel"]
    web = [r for r in per_page if r["surface"] == "Website"]
    app = [r for r in per_page if r["surface"] == "Mobile App"]

    flows = parse_business_flows(FLOWS_DOC)
    bf_main = [r for r in flows if r.get("flow_id", "") <= "BF-09"]
    bf_cross = [r for r in flows if r.get("flow_id", "") >= "BF-10"]

    wb = openpyxl.Workbook()

    # 00 How to use
    ws0 = wb.active
    ws0.title = "00 · How to use"
    ws0.column_dimensions["A"].width = 120
    howto = [
        ("SIMF — Production-Readiness Test Book", TITLE_FONT),
        ("Generated from docs/tests/e2e/*.md + SIMF-Business-Flows.md by tools/testbook/build_testbook.py. Do not hand-edit test rows — re-run the script.", SUB_FONT),
        (f"Catalogue size on {datetime.date.today().isoformat()}: {size['pages']} pages catalogued "
         f"({size['cp']} Control Panel + {size['mobile']} mobile + {size['web']} Website + {size['other']} other), "
         f"{size['scenarios']} Coverage-matrix scenarios. Re-derived by the generator every run — "
         f"quote this line or the roll-up in docs/tests/e2e/README.md, never a number copied into a paragraph.", SUB_FONT),
        ("", None),
        ("How to run", Font(bold=True, size=13)),
        ("1. Bring up the local stack (Test-Guide.md §12.2): API :5175, Control Panel :5158, Website :5115; wait for GET /health → 200.", None),
        ("2. Sign in as the seeded super-admin (password + TOTP via the Get-Totp helper). Never write a literal secret into this workbook.", None),
        ("3. Work sheet by sheet. For each row: perform the Test, then fill Comment / Time (min) / Status / Developer comment / Evidence.", None),
        ("4. Business Flows (sheet 02) and Cross-cutting (sheet 06) carry the full Gherkin steps inline. Per-page sheets (03/04/05) link to the canonical .md in the 'Catalogue ref' column for the full steps.", None),
        ("5. Log every bug on sheet 07 · Defect log with a severity, root cause and the fix commit.", None),
        ("", None),
        ("Status legend", Font(bold=True, size=13)),
        ("Not Run — not yet executed.   Pass — matched the expected result.   Fail — deviated (log a defect).   Blocked — could not run (dependency/environment).   N/A — not applicable to this build (e.g. a deferred feature).", None),
        ("", None),
        ("Priority", Font(bold=True, size=13)),
        ("P0 — must pass to ship (golden paths, auth/permission gates). P1 — must pass for a release. P2 — resilience/edge; exercised once per release window.", None),
        ("", None),
        ("Scope note", Font(bold=True, size=13)),
        ("Deferred features (attendee-facing geofence check-in; exact statistics metric list; dynamic session-category list) and provider stubs (live-video; AI question-filter) are marked N/A, not Fail. See SIMF-Production-Readiness-Round1.md §2.2 and §7.", None),
    ]
    for i, (txt, font) in enumerate(howto, start=1):
        c = ws0.cell(row=i, column=1, value=txt)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            c.font = font

    # test sheets
    ws_bf = wb.create_sheet("02 · Business Flows")
    n_bf = write_rows(ws_bf, bf_main)
    ws_cp = wb.create_sheet("03 · Control Panel")
    n_cp = write_rows(ws_cp, cp)
    ws_web = wb.create_sheet("04 · Website")
    n_web = write_rows(ws_web, web)
    ws_app = wb.create_sheet("05 · Mobile App")
    n_app = write_rows(ws_app, app)
    ws_cross = wb.create_sheet("06 · Cross-cutting")
    n_cross = write_rows(ws_cross, bf_cross)

    # 07 Defect log
    ws_def = wb.create_sheet("07 · Defect log")
    defcols = [
        ("Defect ID", 12), ("Date", 12), ("Source scenario / flow ID", 24),
        ("Surface", 12), ("Severity", 12), ("Title", 40), ("Steps to reproduce", 50),
        ("Expected", 30), ("Actual", 30), ("Root cause", 40), ("Fix commit", 18),
        ("Regression test", 30), ("Status", 12),
    ]
    for c, (name, w) in enumerate(defcols, start=1):
        cell = ws_def.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws_def.column_dimensions[get_column_letter(c)].width = w
    ws_def.freeze_panes = "A2"
    sev = DataValidation(type="list", formula1='"P0-Critical,P1-High,P2-Medium,P3-Low"', allow_blank=True)
    sev.add("E2:E500")
    ws_def.add_data_validation(sev)
    dstat = DataValidation(type="list", formula1='"Open,In progress,Fixed,Verified,Won\'t fix,Deferred"', allow_blank=True)
    dstat.add("M2:M500")
    ws_def.add_data_validation(dstat)

    # 01 Summary (formulas over each sheet's Status column)
    ws_sum = wb.create_sheet("01 · Summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.cell(row=1, column=1, value="SIMF Production-Readiness — run summary").font = TITLE_FONT
    ws_sum.cell(row=2, column=1, value="Live roll-up over the Status columns. Updates as the tester fills each sheet.").font = SUB_FONT
    sheets = [
        ("02 · Business Flows", n_bf),
        ("03 · Control Panel", n_cp),
        ("04 · Website", n_web),
        ("05 · Mobile App", n_app),
        ("06 · Cross-cutting", n_cross),
    ]
    heads = ["Sheet", "Total", "Not Run", "Pass", "Fail", "Blocked", "N/A", "% Pass"]
    hr = 4
    for c, h in enumerate(heads, start=1):
        cell = ws_sum.cell(row=hr, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws_sum.column_dimensions[get_column_letter(c)].width = 20 if c == 1 else 12
    r = hr + 1
    for name, last in sheets:
        rng = f"'{name}'!{STATUS_LETTER}2:{STATUS_LETTER}{max(last,2)}"
        total = max(last - 1, 0)
        ws_sum.cell(row=r, column=1, value=name)
        ws_sum.cell(row=r, column=2, value=total)
        for c, label in ((3, "Not Run"), (4, "Pass"), (5, "Fail"), (6, "Blocked"), (7, "N/A")):
            ws_sum.cell(row=r, column=c, value=f'=COUNTIF({rng},"{label}")')
        ws_sum.cell(row=r, column=8, value=f'=IFERROR(D{r}/(B{r}-G{r}),0)').number_format = "0%"
        r += 1
    # totals row
    ws_sum.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 8):
        L = get_column_letter(c)
        ws_sum.cell(row=r, column=c, value=f"=SUM({L}{hr+1}:{L}{r-1})").font = Font(bold=True)
    ws_sum.cell(row=r, column=8, value=f'=IFERROR(D{r}/(B{r}-G{r}),0)').number_format = "0%"

    # Catalogue size — emitted so no document has to hand-copy it. The TOTAL row
    # above counts workbook ROWS (business flows included, tolerant harvest); this
    # block is the catalogue itself, on the same rule the build gate enforces.
    r += 2
    ws_sum.cell(row=r, column=1, value="Catalogue size (re-derived by the generator)").font = Font(bold=True, size=13)
    r += 1
    for label, value in (
        ("Pages catalogued", size["pages"]),
        ("  · Control Panel", size["cp"]),
        ("  · Mobile App", size["mobile"]),
        ("  · Website", size["web"]),
        ("  · Other", size["other"]),
        ("Coverage-matrix scenarios", size["scenarios"]),
    ):
        ws_sum.cell(row=r, column=1, value=label)
        ws_sum.cell(row=r, column=2, value=value)
        r += 1
    ws_sum.cell(
        row=r, column=1,
        value="Source: docs/tests/e2e/*.md, counted on the same rule as "
              "E2eCatalogueIntegrityTests.The_index_roll_up_matches_the_catalogue_it_describes. "
              "Re-run tools/testbook/build_testbook.py to refresh.",
    ).font = SUB_FONT

    # order the sheets
    order = ["00 · How to use", "01 · Summary", "02 · Business Flows",
             "03 · Control Panel", "04 · Website", "05 · Mobile App",
             "06 · Cross-cutting", "07 · Defect log"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  business flows (BF-01..09): {n_bf}")
    print(f"  control panel:              {n_cp}")
    print(f"  website:                    {n_web}")
    print(f"  mobile app:                 {n_app}")
    print(f"  cross-cutting (BF-10..15):  {n_cross}")
    print(f"  TOTAL test rows:            {n_bf + n_cp + n_web + n_app + n_cross}")
    if not flows:
        print("  NOTE: SIMF-Business-Flows.md not found yet — flow sheets are empty; "
              "re-run after it is authored.", file=sys.stderr)
    report_catalogue_size(size)


if __name__ == "__main__":
    build()
